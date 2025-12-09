from openai import OpenAI
import re
import os

from openpyxl import load_workbook

def remove_think_tags(text: str) -> str:
    return re.sub(r"<think>.*?</think>\s*Reflections:\s*", "", text, flags=re.DOTALL)

prefix = 'The following is an interview with a preparatory student about their learning experiences after taking advanced mathematics courses through traditional face-to-face classroom instruction.\n'

edu_psy = 'Suppose the role of an educational psychology expert and analyze the above observation data about a student interview transcript. Please write reflections to identify motivational triggers and emotional states that correlate with persistent or volatile learning behaviors about the student, studying the psychological factors involved in the learning process. The reflections should make sense given the depth of the interview content above. Only the state of the learners needs to be analyzed.\n'

lea_sci = 'Suppose the role of a Learning Science Expert and analyze the above observation data about a student interview transcript. Please analyze metacognitive strategies and learning patterns to predict the learning behaviors that the student might adopt in actual scenarios, reflecting the emphasis of learning science on understanding and improving the processes and methods of learning. The reflections should make sense given the depth of the interview content above. Only the state of the learners needs to be analyzed.\n'

edu_tech = 'Suppose the role of an Educational Technology Expert and analyze the above observation data about a student interview transcript. Please analyze the educational theories and educational technologies implied therein for learning behavior simulation about the student, concentrating on the application of technical tools to enhance educational practices. The reflections should make sense given the depth of the interview content above. Only the state of the learners needs to be analyzed.\n'

ass_mea = 'Suppose the role of an Assessment Measurement Expert and analyze the above observation data about a student interview transcript. Please operationalize qualitative interview data into Learning status for learning behavior simulation about the student, mirroring the specialized focus of this field on designing effective evaluation methods. The reflections should make sense given the depth of the interview content above. Only the state of the learners needs to be analyzed.\n'

soc_edu = 'Suppose the role of a Sociocultural Education Expert and analyze the above observation data about a student interview transcript. Please analyze the interview by applying sociological principles and methods, emphasizing the role of social and cultural influences on the learning process. The reflections should make sense given the depth of the interview content above. Only the state of the learners needs to be analyzed.'

trailer = 'Output format：\n Reflections：a paragraph containing at most 40 words"'

experts = [edu_psy, lea_sci, edu_tech, ass_mea, soc_edu]
outfile_name = ['edu_psy.txt', 'lea_sci.txt', 'edu_tech.txt', 'ass_mea.txt', 'soc_edu.txt']
    
client = OpenAI(
    base_url="http://222.195.133.39:8000/v1",
    api_key="token-abc123",
)

file_path = "q.xlsx"

wb = load_workbook(file_path)
sheet = wb.active  # 默认第一个工作表


# # 读取第一行作为表头
headers = [sheet.cell(row=1, column=col).value for col in range(16, 29)]

# 遍历后续每一行
for row in sheet.iter_rows(min_row=2, min_col=16, max_col=30, values_only=False):
    row_number = row[0].row 
    values = [cell.value for cell in row]
    os.makedirs(str(row_number), exist_ok=True)

    row_dict = dict(zip(headers, values))  # 每行转为 {列名: 值}
    for key in row_dict:
        q_id = re.match(r"(\d+)、", key).group(1)
        res_path = os.path.join(str(row_number), str(q_id))
        # print(key, row_dict[key])
        os.makedirs(res_path, exist_ok=True)
        for index, item in enumerate(experts):
            prompt = prefix + '问：' + key + '答：' + row_dict[key] + '\n' + item + trailer
            print(prompt)
            completion = client.chat.completions.create(
              model="Qwen/Qwen3-8B",
              messages=[
                {"role": "user", "content": prompt}
              ]
            )
            return_content = completion.choices[0].message.content
            print(return_content)
            clean_content = remove_think_tags(return_content)
            file_path = os.path.join(res_path, outfile_name[index])
            with open(file_path, 'w') as f:
                f.write(clean_content) 

            # interv = '问：' + key + '答：' + row_dict[key]
            # file_path = os.path.join(res_path, 'interview.txt')
            # with open(file_path, 'w') as f:
            #     f.write(interv) 



# "NousResearch/Meta-Llama-3-8B-Instruct"
# Shanghai_AI_Laboratory/internlm3-8b-instruct
# Qwen/Qwen3-8B
# OpenBMB/MiniCPM4-8B